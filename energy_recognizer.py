#!/usr/bin/env python3
"""
Программа для распознавания банок энергетиков через камеру.
Определяет производителя и вкус, сохраняет фото и данные в Excel.
"""

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import tkinter as tk
from tkinter import messagebox, filedialog
from PIL import Image, ImageTk

# Конфигурация
EXCEL_FILE = "energy_drinks.xlsx"
PHOTOS_DIR = "photos"

# База данных цветов/логотипов для простых энергетиков (упрощенная версия)
# В реальном проекте тут была бы нейросеть или более сложная логика
BRAND_TEMPLATES = {
    "Red Bull": {
        "colors": [(200, 50, 50), (255, 255, 0)],  # красный, желтый
        "flavors": {
            "original": [(200, 50, 50), (255, 255, 0)],
            "sugarfree": [(200, 200, 200), (0, 100, 200)],
            "blueberry": [(0, 0, 200), (200, 50, 50)]
        }
    },
    "Monster": {
        "colors": [(0, 0, 0), (0, 255, 0)],  # черный, зеленый
        "flavors": {
            "original": [(0, 0, 0), (0, 255, 0)],
            "ultra": [(255, 255, 255), (0, 0, 0)],
            "pipeline": [(255, 165, 0), (0, 0, 0)]
        }
    },
    "Adrenaline Rush": {
        "colors": [(0, 0, 0), (255, 215, 0)],  # черный, золотой
        "flavors": {
            "original": [(0, 0, 0), (255, 215, 0)],
            "citrus": [(255, 165, 0), (0, 0, 0)]
        }
    },
    "Gorilla": {
        "colors": [(255, 0, 0), (0, 0, 0)],
        "flavors": {
            "original": [(255, 0, 0), (0, 0, 0)],
            "tropical": [(255, 165, 0), (0, 255, 0)]
        }
    },
    "Flash Up": {
        "colors": [(255, 255, 0), (0, 0, 0)],
        "flavors": {
            "original": [(255, 255, 0), (0, 0, 0)],
            "zero": [(255, 255, 255), (0, 0, 0)]
        }
    }
}

class EnergyDrinkRecognizer:
    def __init__(self):
        self.root = None
        self.cap = None
        self.video_label = None
        self.current_frame = None
        self.detected_brand = None
        self.detected_flavor = None
        self.photos_dir = PHOTOS_DIR
        
        if not os.path.exists(self.photos_dir):
            os.makedirs(self.photos_dir)
        
        self.init_excel()
    
    def init_excel(self):
        """Инициализация Excel файла"""
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Energy Drinks"
            ws.append(["ID", "Дата", "Производитель", "Вкус", "Фото", "Статус"])
            wb.save(EXCEL_FILE)
            print(f"Создан файл {EXCEL_FILE}")
    
    def analyze_colors(self, frame):
        """Анализ цветов на изображении для определения бренда и вкуса"""
        # Уменьшаем изображение для скорости
        small_frame = cv2.resize(frame, (100, 150))
        hsv = cv2.cvtColor(small_frame, cv2.COLOR_BGR2HSV)
        
        # Получаем средние цвета из разных частей изображения
        h, w, _ = small_frame.shape
        center_region = small_frame[h//4:3*h//4, w//4:3*w//4]
        avg_color = np.mean(center_region, axis=(0, 1))
        
        best_brand = None
        best_flavor = None
        best_score = 0
        
        for brand, data in BRAND_TEMPLATES.items():
            template_colors = data["colors"]
            
            # Простая проверка совпадения цветов
            score = 0
            for tc in template_colors:
                # Проверяем наличие похожих цветов
                lower = np.array([max(0, c - 40) for c in tc])
                upper = np.array([min(255, c + 40) for c in tc])
                mask = cv2.inRange(small_frame, lower, upper)
                color_pixels = cv2.countNonZero(mask)
                score += color_pixels
            
            if score > best_score:
                best_score = score
                best_brand = brand
                
                # Определяем вкус
                best_flavor_score = 0
                for flavor, fcolors in data["flavors"].items():
                    fscore = 0
                    for fc in fcolors:
                        lower = np.array([max(0, c - 30) for c in fc])
                        upper = np.array([min(255, c + 30) for c in fc])
                        mask = cv2.inRange(small_frame, lower, upper)
                        fscore += cv2.countNonZero(mask)
                    
                    if fscore > best_flavor_score:
                        best_flavor_score = fscore
                        best_flavor = flavor
        
        # Если уверенность низкая, возвращем "Неизвестно"
        if best_score < 5000:  # порог уверенности
            return "Неизвестно", "Неизвестно"
        
        return best_brand, best_flavor if best_flavor else "original"
    
    def check_if_exists(self, brand, flavor):
        """Проверка наличия такой банки в Excel"""
        if not os.path.exists(EXCEL_FILE):
            return False
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == brand and row[3] == flavor:
                return True
        
        return False
    
    def save_to_excel(self, brand, flavor, photo_path):
        """Сохранение данных в Excel"""
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Проверка на дубликат
        exists = self.check_if_exists(brand, flavor)
        
        next_id = 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and isinstance(row[0], int) and row[0] >= next_id:
                next_id = row[0] + 1
        
        status = "Уже имеется" if exists else "Новая"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([next_id, timestamp, brand, flavor, photo_path, status])
        wb.save(EXCEL_FILE)
        
        return exists
    
    def capture_and_save(self):
        """Захват кадра и сохранение"""
        if self.current_frame is None or self.detected_brand is None:
            messagebox.showwarning("Внимание", "Сначала поднесите банку к камере!")
            return
        
        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{self.detected_brand}_{self.detected_flavor}.jpg"
        photo_path = os.path.join(self.photos_dir, filename)
        
        # Сохраняем фото
        cv2.imwrite(photo_path, self.current_frame)
        
        # Сохраняем в Excel
        exists = self.save_to_excel(self.detected_brand, self.detected_flavor, photo_path)
        
        if exists:
            messagebox.showinfo("Результат", f"Такая банка уже имеется в базе!\n\nПроизводитель: {self.detected_brand}\nВкус: {self.detected_flavor}")
        else:
            messagebox.showinfo("Результат", f"Новая банка добавлена!\n\nПроизводитель: {self.detected_brand}\nВкус: {self.detected_flavor}\nФото: {photo_path}")
    
    def update_video(self):
        """Обновление видео с камеры"""
        ret, frame = self.cap.read()
        if ret:
            self.current_frame = frame
            
            # Анализ
            brand, flavor = self.analyze_colors(frame)
            self.detected_brand = brand
            self.detected_flavor = flavor
            
            # Отображение результатов на кадре
            info_text = f"Бренд: {brand} | Вкус: {flavor}"
            cv2.putText(frame, info_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            # Преобразование для Tkinter
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame_rgb)
            imgtk = ImageTk.PhotoImage(image=img)
            
            self.video_label.imgtk = imgtk
            self.video_label.configure(image=imgtk)
        
        self.root.after(30, self.update_video)
    
    def open_excel(self):
        """Открытие Excel файла"""
        if os.path.exists(EXCEL_FILE):
            try:
                os.startfile(EXCEL_FILE)
            except:
                # Для Linux/Mac
                import subprocess
                try:
                    subprocess.call(['xdg-open', EXCEL_FILE])
                except:
                    subprocess.call(['open', EXCEL_FILE])
        else:
            messagebox.showwarning("Файл не найден", f"Файл {EXCEL_FILE} еще не создан.")
    
    def create_gui(self):
        """Создание графического интерфейса"""
        self.root = tk.Tk()
        self.root.title("Распознавание энергетиков")
        self.root.geometry("800x700")
        
        # Заголовок
        title_label = tk.Label(self.root, text="Поднесите банку энергетика к камере", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # Видео
        self.video_label = tk.Label(self.root)
        self.video_label.pack(pady=10)
        
        # Информация
        self.info_label = tk.Label(self.root, text="", font=("Arial", 12))
        self.info_label.pack(pady=5)
        
        # Кнопки
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=20)
        
        capture_btn = tk.Button(btn_frame, text="📸 Сделать фото и сохранить", command=self.capture_and_save, 
                               font=("Arial", 12), bg="#4CAF50", fg="white", padx=20, pady=10)
        capture_btn.pack(side=tk.LEFT, padx=10)
        
        excel_btn = tk.Button(btn_frame, text="📊 Открыть Excel", command=self.open_excel,
                             font=("Arial", 12), bg="#2196F3", fg="white", padx=20, pady=10)
        excel_btn.pack(side=tk.LEFT, padx=10)
        
        # Инструкция
        instr_label = tk.Label(self.root, text="Инструкция:\n1. Поднесите банку к камере\n2. Дождитесь определения бренда и вкуса\n3. Нажмите 'Сделать фото'\n4. Данные сохранятся в Excel", 
                              justify=tk.LEFT, font=("Arial", 10))
        instr_label.pack(pady=20)
        
        # Старт камеры
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showerror("Ошибка", "Не удалось открыть камеру!")
            self.root.destroy()
            return
        
        self.update_video()
        self.root.mainloop()
        
        self.cap.release()
        cv2.destroyAllWindows()

def main():
    print("Запуск программы распознавания энергетиков...")
    recognizer = EnergyDrinkRecognizer()
    recognizer.create_gui()

if __name__ == "__main__":
    main()
