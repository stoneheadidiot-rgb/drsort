#!/usr/bin/env python3
"""
Программа для распознавания банок энергетиков через камеру с использованием YOLO.
Определяет производителя и вкус, сохраняет фото и данные в Excel.
Аналогично Google Поиск по картинке - использует нейросеть для распознавания.
"""

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from ultralytics import YOLO

# Конфигурация
EXCEL_FILE = "energy_drinks.xlsx"
PHOTOS_DIR = "photos"
YOLO_MODEL = "yolo11n.pt"  # Легкая и быстрая модель YOLO11

# Ключевые слова для определения брендов энергетиков
BRAND_KEYWORDS = {
    "Red Bull": ["red bull", "redbull"],
    "Monster": ["monster"],
    "Adrenaline Rush": ["adrenaline", "adrenaline rush"],
    "Burn": ["burn"],
    "Rockstar": ["rockstar", "rock star"],
    "Gorilla": ["gorilla"],
    "Flash Up": ["flash", "flash up"],
    "Hell Energy": ["hell", "hell energy"],
    "Drive": ["drive"],
    "Volt": ["volt"],
    "Power Horse": ["power horse", "powerhorse"],
    "Black Monster": ["black monster"]
}

# Ключевые слова для определения вкуса
FLAVOR_KEYWORDS = {
    "Original": ["original", "classic", "regular", "the original"],
    "Sugar Free": ["sugarfree", "sugar free", "zero sugar", "zero", "light", "no sugar"],
    "Tropical": ["tropical", "fruit", "mango", "pineapple", "passion fruit"],
    "Berry": ["berry", "blueberry", "raspberry", "strawberry", "mixed berry"],
    "Citrus": ["citrus", "lemon", "lime", "orange", "grapefruit"],
    "Apple": ["apple", "green apple"],
    "Watermelon": ["watermelon", "melon"],
    "Cherry": ["cherry"],
    "Banana": ["banana"],
    "Coffee": ["coffee", "mocha", "latte", "cappuccino"],
    "Energy": ["energy"],
    "Blue Edition": ["blue", "blue edition"],
    "Red Edition": ["red", "red edition"],
    "Green Edition": ["green", "green edition"],
    "Yellow Edition": ["yellow", "yellow edition"],
    "White Edition": ["white", "white edition"],
    "Pipeline Punch": ["pipeline", "pipeline punch"],
    "Ultra Blue": ["ultra blue", "ultra"],
    "Mean Bean": ["mean bean", "java"],
    "Mango Loco": ["mango loco", "loco"]
}


class EnergyRecognizerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Raspoznavanie energetikov (YOLO AI)")
        self.root.geometry("900x750")
        
        self.cap = None
        self.current_frame = None
        self.detected_brand = None
        self.detected_flavor = None
        self.detected_confidence = 0
        self.model = None
        self.is_model_loaded = False
        
        if not os.path.exists(PHOTOS_DIR):
            os.makedirs(PHOTOS_DIR)
        
        self.setup_ui()
        self.load_model_async()
    
    def setup_ui(self):
        title_label = tk.Label(self.root, text="Podnesite banku energetika k kamere", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        video_frame = tk.Frame(self.root)
        video_frame.pack(pady=10)
        
        self.video_label = tk.Label(video_frame)
        self.video_label.pack()
        
        self.status_label = tk.Label(self.root, text="Zagruzka neyroseti...", 
                                    font=("Arial", 12), fg="orange")
        self.status_label.pack(pady=5)
        
        self.confidence_label = tk.Label(self.root, text="", font=("Arial", 10))
        self.confidence_label.pack(pady=2)
        
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=15)
        
        self.capture_btn = tk.Button(button_frame, text="Sdelat foto i sokhranit", 
                                    command=self.capture_and_save, state=tk.DISABLED,
                                    font=("Arial", 12), bg="#4CAF50", fg="white", 
                                    padx=20, pady=10)
        self.capture_btn.pack(side=tk.LEFT, padx=10)
        
        excel_btn = tk.Button(button_frame, text="Otkryt Excel", 
                             command=self.open_excel,
                             font=("Arial", 12), bg="#2196F3", fg="white", 
                             padx=20, pady=10)
        excel_btn.pack(side=tk.LEFT, padx=10)
        
        quit_btn = tk.Button(button_frame, text="Vykhod", command=self.quit_app,
                            font=("Arial", 12), bg="#f44336", fg="white", 
                            padx=20, pady=10)
        quit_btn.pack(side=tk.LEFT, padx=10)
        
        instr_text = """Instruktsiya:
1. Dozhdives zagruzki modeli (neskolko sekund)
2. Podnesite banku energetika k kamere
3. Neyroset avtomaticheski opredelit brend i vkus
4. Nazhmite Sdelat foto dlya sokhraneniya v bazu
5. Esli takaya banka uzhe est - poluchite uvedomlenie"""
        
        instr_label = tk.Label(self.root, text=instr_text, justify=tk.LEFT, 
                              font=("Arial", 10), fg="gray")
        instr_label.pack(pady=10)
    
    def load_model_async(self):
        self.root.after(100, self._load_model)
    
    def _load_model(self):
        try:
            self.status_label.config(text="Zagruzka YOLO modeli...")
            self.root.update()
            
            self.model = YOLO(YOLO_MODEL)
            
            self.is_model_loaded = True
            self.status_label.config(text="Model zagruzhena! Navedite kameru na banku.", 
                                    fg="green")
            
            self.start_camera()
            
        except Exception as e:
            messagebox.showerror("Oshibka", f"Ne udalos zagruzit model:\n{e}")
            self.status_label.config(text="Oshibka zagruzki modeli", fg="red")
    
    def start_camera(self):
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showerror("Oshibka", "Ne udalos otkryt kameru")
            self.root.quit()
            return
        
        self.update_frame()
    
    def detect_energy_drink(self, frame):
        if self.model is None or not self.is_model_loaded:
            return None, None, 0, None
        
        results = self.model(frame, verbose=False, conf=0.25, iou=0.5)
        
        best_brand = None
        best_flavor = None
        best_confidence = 0
        best_box = None
        
        for result in results:
            boxes = result.boxes
            if boxes is None:
                continue
            
            for box in boxes:
                cls = int(box.cls[0])
                conf = float(box.conf[0])
                
                class_name = result.names[cls].lower()
                
                if any(keyword in class_name for keyword in ["can", "bottle", "cup", "drink", "beverage"]):
                    brand = "Generic Can"
                    
                    for b_name, keywords in BRAND_KEYWORDS.items():
                        if any(kw in class_name for kw in keywords):
                            brand = b_name
                            break
                    
                    flavor = self.infer_flavor(class_name)
                    
                    if conf > best_confidence:
                        best_confidence = conf
                        best_brand = brand
                        best_flavor = flavor
                        best_box = box.xyxy[0]
        
        return best_brand, best_flavor, best_confidence, best_box
    
    def infer_flavor(self, class_name):
        class_lower = class_name.lower()
        
        for flavor, keywords in FLAVOR_KEYWORDS.items():
            if any(kw in class_lower for kw in keywords):
                return flavor
        
        return "Unknown"
    
    def update_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            self.root.after(100, self.update_frame)
            return
        
        self.current_frame = frame.copy()
        
        self.detected_brand, self.detected_flavor, confidence, box = self.detect_energy_drink(frame)
        
        display_frame = frame.copy()
        
        if self.detected_brand and confidence > 0.3:
            status_text = f"Brend: {self.detected_brand}"
            if self.detected_flavor and self.detected_flavor != "Unknown":
                status_text += f" | Vkus: {self.detected_flavor}"
            status_text += f" (uverennost: {confidence:.1%})"
            
            self.status_label.config(text=status_text, fg="green")
            self.confidence_label.config(text=f"Confidence: {confidence:.3f}")
            self.capture_btn.config(state=tk.NORMAL)
            
            if box is not None:
                x1, y1, x2, y2 = map(int, box)
                cv2.rectangle(display_frame, (x1, y1), (x2, y2), (0, 255, 0), 3)
                label = f"{self.detected_brand}"
                if self.detected_flavor and self.detected_flavor != "Unknown":
                    label += f" - {self.detected_flavor}"
                cv2.putText(display_frame, label, (x1, y1 - 10),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        else:
            self.status_label.config(text="Energetik ne raspoznan. Podnesite banku blizhe.", fg="red")
            self.confidence_label.config(text="")
            self.capture_btn.config(state=tk.DISABLED)
        
        rgb_frame = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb_frame)
        imgtk = ImageTk.PhotoImage(image=img)
        
        self.video_label.imgtk = imgtk
        self.video_label.config(image=imgtk)
        
        self.root.after(100, self.update_frame)
    
    def capture_and_save(self):
        if self.current_frame is None or self.detected_brand is None:
            messagebox.showwarning("Preduprezhdenie", "Net dannykh dlya sokhraneniya")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        brand_safe = self.detected_brand.replace(" ", "_")
        flavor_safe = (self.detected_flavor or "Unknown").replace(" ", "_")
        
        photo_filename = f"{timestamp}_{brand_safe}_{flavor_safe}.jpg"
        photo_path = os.path.join(PHOTOS_DIR, photo_filename)
        
        cv2.imwrite(photo_path, self.current_frame)
        
        wb = self.create_or_load_excel()
        is_duplicate = self.check_if_exists(wb, self.detected_brand, self.detected_flavor)
        
        ws = wb.active
        next_id = self.get_next_id(ws)
        
        status = "Uzhe imeetsya" if is_duplicate else "Novaya"
        
        ws.append([
            next_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            self.detected_brand,
            self.detected_flavor or "Ne opredelen",
            photo_path,
            status
        ])
        
        wb.save(EXCEL_FILE)
        
        if is_duplicate:
            messagebox.showinfo("Rezultat", 
                               f"Takaya banka uzhe est v baze!\n\n"
                               f"Proizvoditel: {self.detected_brand}\n"
                               f"Vkus: {self.detected_flavor or 'Ne opredelen'}\n"
                               f"Foto sokhraneno.")
        else:
            messagebox.showinfo("Rezultat", 
                               f"Novaya banka dobavlena!\n\n"
                               f"Proizvoditel: {self.detected_brand}\n"
                               f"Vkus: {self.detected_flavor or 'Ne opredelen'}\n"
                               f"Foto: {photo_path}")
    
    def create_or_load_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Energy Drinks"
            ws.append(["ID", "Data", "Proizvoditel", "Vkus", "Foto", "Status"])
            wb.save(EXCEL_FILE)
        return load_workbook(EXCEL_FILE)
    
    def get_next_id(self, ws):
        max_id = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and isinstance(row[0], (int, float)):
                if row[0] > max_id:
                    max_id = int(row[0])
        return max_id + 1
    
    def check_if_exists(self, wb, brand, flavor):
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == brand:
                    row_flavor = row[3]
                    if flavor is None or flavor == "Unknown":
                        if row_flavor in [None, "Unknown", "Ne opredelen"]:
                            return True
                    elif row_flavor == flavor:
                        return True
        except Exception:
            pass
        return False
    
    def open_excel(self):
        if not os.path.exists(EXCEL_FILE):
            self.create_or_load_excel()
        
        try:
            if os.name == 'nt':
                os.startfile(EXCEL_FILE)
            elif os.name == 'posix':
                os.system(f"open {EXCEL_FILE}")
            else:
                os.system(f"xdg-open {EXCEL_FILE}")
        except Exception as e:
            messagebox.showerror("Oshibka", f"Ne udalos otkryt Excel: {e}")
    
    def quit_app(self):
        if self.cap:
            self.cap.release()
        self.root.quit()
        self.root.destroy()


def main():
    print("=" * 50)
    print("Programma raspoznavaniya energetikov (YOLO AI)")
    print("=" * 50)
    print("Zapusk graficheskogo interfeysa...")
    print("Pervyy zapusk mozhet zanyat vremya na zagruzku modeli.")
    print()
    
    root = tk.Tk()
    app = EnergyRecognizerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
